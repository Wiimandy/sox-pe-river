import os
import sys
import json
import urllib.request
import datetime
import pandas as pd
import yfinance as yf

# Tickers are fetched dynamically from Cathay API (fallback list is in fetch_yfinance_bottomup)

CSV_FILE = "sox_daily_tracking_log.csv"
XLSX_FILE = "sox_daily_tracking_log.xlsx"

def save_to_excel_preserve_sheets(df_log, xlsx_file):
    import openpyxl
    if os.path.exists(xlsx_file):
        print(f"[Excel] Updating sheet 'Daily_Log' in existing file: {xlsx_file}")
        book = openpyxl.load_workbook(xlsx_file)
        if "Daily_Log" in book.sheetnames:
            std = book["Daily_Log"]
            book.remove(std)
        sheet = book.create_sheet("Daily_Log", 0)
        
        # Write headers
        headers = list(df_log.columns)
        sheet.append(headers)
        
        # Write rows
        for row in df_log.itertuples(index=False):
            clean_row = []
            for val in row:
                if pd.isna(val):
                    clean_row.append(None)
                elif isinstance(val, (datetime.date, datetime.datetime)):
                    clean_row.append(val.strftime('%Y-%m-%d'))
                else:
                    clean_row.append(val)
            sheet.append(clean_row)
            
        book.save(xlsx_file)
        book.close()
        print(f"[Excel] Successfully saved log sheet to {xlsx_file}")
    else:
        print(f"[Excel] Creating new Excel file: {xlsx_file}")
        metadata_data = [
            {"項目 (Item)": "專案名稱 (Project)", "內容 (Content)": "SOX Index P/E River Dashboard (費城半導體本益比河流圖及每日追蹤日誌)"},
            {"項目 (Item)": "開發歷程主要分期 (Development Phases)", "內容 (Content)": "分為以下兩個階段："},
            {"項目 (Item)": "  階段一：等股數相加估算期 (Phase 1: Simple Sum)", "內容 (Content)": "2026-06-29 及以前。本益比計算採用「成份股股價加總 / EPS加總」，權重等同於每股 1 股。"},
            {"項目 (Item)": "  階段二：動態加權本益比期 (Phase 2: Weighted)", "內容 (Content)": "2026-06-30 起。串接國泰美國費城半導體 ETF (00830) 官網成份股及權重 API，採動態加權倒數法（加權盈餘殖利率倒數）計算本益比。股價基準對齊真實費半指數 (^SOX)。"},
            {"項目 (Item)": "數據基準對齊資訊 (Data Reference)", "內容 (Content)": "真實費半指數收盤價 (^SOX) 及成分股週頻/日頻 historical/forward EPS"},
            {"項目 (Item)": "國泰 00830 權重數據 API 來源", "內容 (Content)": "https://cwapi.cathaysite.com.tw/api/ETF/GetIndexStockWeights?FundCode=BO&status=1"},
            {"項目 (Item)": "iShares SOXX 基金指標數據來源", "內容 (Content)": "https://www.ishares.com/us/products/239705/ishares-phlx-semiconductor-etf (擷取 Trailing P/E)"},
            {"項目 (Item)": "更新記錄時間 (Update Cron)", "內容 (Content)": "每天早上 10:00 自動執行 daily update & publish"},
        ]
        df_meta = pd.DataFrame(metadata_data)
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            df_log.to_excel(writer, sheet_name="Daily_Log", index=False)
            df_meta.to_excel(writer, sheet_name="Metadata_DevNotes", index=False)
        print(f"[Excel] Successfully created Excel workbook: {xlsx_file}")

# Other indices to track with Rolling 12M Forward PE (simple LSEG-only log)
OTHER_INDICES = [
    {"ticker": ".SPX",  "name": "SPX",  "csv": "spx_daily_tracking_log.csv"},
    {"ticker": ".IXIC", "name": "IXIC", "csv": "ixic_daily_tracking_log.csv"},
    {"ticker": ".DJI",  "name": "DJI",  "csv": "dji_daily_tracking_log.csv"},
]

def fetch_lseg_data():
    """Fetch latest price, PE, and Forward PE for .SOX index from LSEG Workspace."""
    print("[LSEG] Attempting to connect via LSEG Data Library...")
    import lseg.data as ld
    
    opened = False
    try:
        ld.open_session()
        opened = True
        print("[LSEG] Session opened successfully.")
    except Exception as e:
        print(f"[LSEG] Direct session failed: {e}. Trying desktop.workspace...")
        try:
            ld.open_session(name="desktop.workspace")
            opened = True
            print("[LSEG] Desktop workspace session opened.")
        except Exception as e2:
            print(f"[LSEG] Session open failed: {e2}")
            
    if not opened:
        # Fallback to Eikon
        print("[LSEG] Attempting fallback to Eikon Data API...")
        try:
            import eikon as ek
            app_key = os.environ.get("LSEG_APP_KEY", "dummy_app_key")
            ek.set_app_key(app_key)
            df, err = ek.get_data(
                instruments=[".SOX"],
                fields=["TR.PriceClose", "TR.Index_PE_RTRS", "TR.Index_EST_PE_Y1_RTRS", "TR.Index_EST_PE_Y2_RTRS"]
            )
            if err:
                print(f"[Eikon] API warning: {err}")
            if df is not None and not df.empty:
                # Rename columns explicitly by position
                df.columns = ['Instrument', 'Price Close', 'Calculated PE Ratio', 'Y1 PE Ratio', 'Y2 PE Ratio']
                return df
        except Exception as e3:
            print(f"[Eikon] Fallback failed: {e3}")
            return None
            
    try:
        df = ld.get_data(
            universe=[".SOX"],
            fields=["TR.PriceClose", "TR.Index_PE_RTRS", "TR.Index_EST_PE_Y1_RTRS", "TR.Index_EST_PE_Y2_RTRS"]
        )
        ld.close_session()
        if df is not None and not df.empty:
            df.columns = ['Instrument', 'Price Close', 'Calculated PE Ratio', 'Y1 PE Ratio', 'Y2 PE Ratio']
        return df
    except Exception as e:
        print(f"[LSEG] Query failed: {e}")
        try:
            ld.close_session()
        except:
            pass
        return None


def fetch_cathay_weights():
    print("[Cathay API] Fetching current index weights for 00830 (FundCode=BO)...")
    url = "https://cwapi.cathaysite.com.tw/api/ETF/GetIndexStockWeights?FundCode=BO&status=1"
    req = urllib.request.Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Referer': 'https://www.cathaysite.com.tw/'
    })
    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            content = response.read().decode('utf-8')
        data = json.loads(content)
        if not data.get('success'):
            raise ValueError(f"API returned failure: {data.get('returnMessage')}")
        
        stock_weights = data.get('result', {}).get('stockWeights', [])
        weights = {}
        for sw in stock_weights:
            ticker = sw['stockCode'].replace('.US', '')
            weight = float(sw['weights']) / 100.0
            weights[ticker] = weight
        print(f"[Cathay API] Successfully fetched {len(weights)} components and weights.")
        return weights
    except Exception as e:
        print(f"[Cathay API] Error fetching weights: {e}")
        return None

def fetch_yfinance_bottomup():
    """Fetch prices and EPS for SOX components from yfinance and compute weighted bottom-up PE."""
    weights = fetch_cathay_weights()
    if weights is None:
        print("[yfinance] Warning: Using hardcoded tickers as fallback because Cathay API failed.")
        fallback_tickers = [
            'NVDA', 'AVGO', 'AMD',  'QCOM', 'AMAT',
            'LRCX', 'KLAC', 'MU',   'TXN',  'ADI',
            'MCHP', 'ON',   'MRVL', 'NXPI', 'MPWR',
            'SWKS', 'QRVO', 'TER',  'ENTG', 'STM',
            'ASML', 'INTC', 'MKSI', 'WOLF', 'ACLS',
            'CAMT', 'FORM', 'SITM', 'RMBS', 'COHU'
        ]
        weights = {t: 1.0 / len(fallback_tickers) for t in fallback_tickers}
        
    tickers = list(weights.keys())
    print(f"[yfinance] Fetching data for {len(tickers)} component stocks...")
    yf_index_price = None
    try:
        sox_info = yf.Ticker("^SOX").info
        yf_index_price = sox_info.get("currentPrice") or sox_info.get("regularMarketPrice")
        if yf_index_price is None:
            sox_history = yf.Ticker("^SOX").history(period="5d")
            if sox_history is not None and not sox_history.empty:
                yf_index_price = float(sox_history["Close"].dropna().iloc[-1])
    except Exception as e:
        print(f"[yfinance] Warning: Could not fetch ^SOX index price: {e}")
    
    rows = []
    missing_fwd = []
    missing_trail = []
    
    # Calculate rolling weights based on day of year
    today = datetime.datetime.now()
    day_of_year = today.timetuple().tm_yday
    is_leap = (today.year % 4 == 0 and (today.year % 100 != 0 or today.year % 400 == 0))
    days_in_year = 366 if is_leap else 365
    weight_y2 = day_of_year / days_in_year
    weight_y1 = 1.0 - weight_y2
    
    for tkr in tickers:
        try:
            t = yf.Ticker(tkr)
            info = t.info
            price = info.get('currentPrice') or info.get('regularMarketPrice')
            
            # Fetch 0y and +1y estimates for rolling forward EPS
            df_est = t.earnings_estimate
            fwd_eps = None
            if df_est is not None and not df_est.empty and 'avg' in df_est.columns:
                if '0y' in df_est.index and '+1y' in df_est.index:
                    eps_y1 = df_est.loc['0y', 'avg']
                    eps_y2 = df_est.loc['+1y', 'avg']
                    if pd.notna(eps_y1) and pd.notna(eps_y2):
                        fwd_eps = weight_y1 * eps_y1 + weight_y2 * eps_y2
            
            # Fallback
            if fwd_eps is None:
                fwd_eps = info.get('forwardEps')
                
            trail_eps = info.get('trailingEps')
            
            rows.append({
                'Ticker': tkr,
                'Price': price,
                'Trail_EPS': trail_eps,
                'Fwd_EPS': fwd_eps,
                'Weight': weights[tkr]
            })
            
            if fwd_eps is None:
                missing_fwd.append(tkr)
            if trail_eps is None:
                missing_trail.append(tkr)
        except Exception as e:
            print(f"[yfinance] Error fetching {tkr}: {e}")
            missing_fwd.append(tkr)
            missing_trail.append(tkr)
            
    df = pd.DataFrame(rows).dropna(subset=['Price'])
    if df.empty:
        print("[yfinance] Error: Could not fetch data for any components.")
        return None, None, yf_index_price, [], []
        
    # Trailing PE: weighted harmonic mean
    df_trail = df.dropna(subset=['Trail_EPS'])
    if not df_trail.empty:
        sum_trail_weight = df_trail['Weight'].sum()
        if sum_trail_weight > 0:
            # Re-normalize weights
            df_trail['Norm_Weight'] = df_trail['Weight'] / sum_trail_weight
            sum_trail_yield = (df_trail['Norm_Weight'] * (df_trail['Trail_EPS'] / df_trail['Price'])).sum()
            yf_trail_pe = 1.0 / sum_trail_yield if sum_trail_yield > 0 else None
        else:
            yf_trail_pe = None
    else:
        yf_trail_pe = None
        
    # Forward PE: weighted harmonic mean
    df_fwd = df.dropna(subset=['Fwd_EPS'])
    if not df_fwd.empty:
        sum_fwd_weight = df_fwd['Weight'].sum()
        if sum_fwd_weight > 0:
            df_fwd['Norm_Weight'] = df_fwd['Weight'] / sum_fwd_weight
            sum_fwd_yield = (df_fwd['Norm_Weight'] * (df_fwd['Fwd_EPS'] / df_fwd['Price'])).sum()
            yf_fwd_pe = 1.0 / sum_fwd_yield if sum_fwd_yield > 0 else None
        else:
            yf_fwd_pe = None
    else:
        yf_fwd_pe = None
        
    return yf_trail_pe, yf_fwd_pe, yf_index_price, missing_trail, missing_fwd


def fetch_soxx_pe_web():
    """Fetch latest fund characteristics (PE) from iShares SOXX website."""
    import re
    import html
    print("[iShares] Fetching SOXX ETF stats from iShares website...")
    url = "https://www.ishares.com/us/products/239705/ishares-phlx-semiconductor-etf"
    req = urllib.request.Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    })
    
    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            content = response.read().decode('utf-8')
            
        decoded_text = html.unescape(content)
        
        # Extract P/E Ratio
        pe_match = re.search(r'"label"\s*:\s*"P/E Ratio"\s*,\s*"formattedValue"\s*:\s*"([^"]*?)"', decoded_text)
        pe_val = None
        if pe_match:
            pe_val = float(pe_match.group(1).replace(',', ''))
            
        # Extract as-of date
        date_match = re.search(r'"numHoldings"\s*:\s*\{[^}]*?"formattedAsOfDate"\s*:\s*"([^"]*?)"', decoded_text)
        as_of_date = None
        if date_match:
            as_of_date = date_match.group(1)
        else:
            # Fallback date search
            date_matches = re.findall(r'"formattedAsOfDate"\s*:\s*"([^"]*?)"', decoded_text)
            if date_matches:
                for d in date_matches:
                    if any(m in d for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']):
                        as_of_date = d
                        break
                        
        print(f"[iShares] Scraped successfully. PE={pe_val}, As-of={as_of_date}")
        return as_of_date, pe_val, None  # Forward PE is None for SOXX
    except Exception as e:
        print(f"[iShares] Error fetching SOXX: {e}")
        return None, None, None


def main():
    run_date = datetime.date.today().strftime('%Y-%m-%d')
    today = datetime.date.today()
    print("=" * 70)
    print(f"  SOX Index Daily PE Tracking Log - Run Date: {run_date}")
    print("=" * 70)
    
    # 1. Fetch LSEG
    lseg_price, lseg_pe, lseg_fwd_pe = None, None, None
    try:
        df_lseg = fetch_lseg_data()
        if df_lseg is not None and not df_lseg.empty:
            lseg_price = float(df_lseg['Price Close'].iloc[0])
            lseg_pe = float(df_lseg['Calculated PE Ratio'].iloc[0])
            
            y1_pe = float(df_lseg['Y1 PE Ratio'].iloc[0])
            y2_pe_val = df_lseg['Y2 PE Ratio'].iloc[0]
            if pd.notna(y2_pe_val):
                y2_pe = float(y2_pe_val)
                day_of_year = today.timetuple().tm_yday
                is_leap = (today.year % 4 == 0 and (today.year % 100 != 0 or today.year % 400 == 0))
                days_in_year = 366 if is_leap else 365
                weight_y2 = day_of_year / days_in_year
                weight_y1 = 1.0 - weight_y2
                lseg_fwd_pe = weight_y1 * y1_pe + weight_y2 * y2_pe
            else:
                lseg_fwd_pe = y1_pe
            
            print(f"[LSEG] Close={lseg_price:.2f}, PE={lseg_pe:.2f}x, FwdPE={lseg_fwd_pe:.2f}x (Rolling 12M)")
        else:
            print("[LSEG] Warning: No LSEG data retrieved. Logging as N/A.")
    except Exception as e:
        print(f"[LSEG] Error processing LSEG data: {e}. Logging as N/A.")
        
    # 2. Fetch yfinance
    yf_pe, yf_fwd_pe, yf_index_price = None, None, None
    missing_t, missing_f = [], []
    try:
        yf_pe, yf_fwd_pe, yf_index_price, missing_t, missing_f = fetch_yfinance_bottomup()
        if yf_pe is not None:
            print(f"[yfinance] Calculated PE={yf_pe:.2f}x, FwdPE={yf_fwd_pe:.2f}x")
        else:
            print("[yfinance] Warning: No yfinance PE calculated.")
    except Exception as e:
        print(f"[yfinance] Error during calculations: {e}")
        
    # 3. Fetch SOXX from iShares
    soxx_date, soxx_pe, soxx_fwd_pe = fetch_soxx_pe_web()
    if soxx_pe is not None:
        print(f"[iShares] SOXX PE={soxx_pe:.2f}x (as of {soxx_date})")
        
    # 4. Calculate metrics
    diff_pe_yf = (lseg_pe - yf_pe) if (lseg_pe is not None and yf_pe is not None) else None
    diff_fwd_pe_yf = (lseg_fwd_pe - yf_fwd_pe) if (lseg_fwd_pe is not None and yf_fwd_pe is not None) else None
    ratio_pe_yf = (lseg_pe / yf_pe) if (lseg_pe is not None and yf_pe is not None and yf_pe != 0) else None
    ratio_fwd_pe_yf = (lseg_fwd_pe / yf_fwd_pe) if (lseg_fwd_pe is not None and yf_fwd_pe is not None and yf_fwd_pe != 0) else None
    
    diff_pe_soxx = (lseg_pe - soxx_pe) if (lseg_pe is not None and soxx_pe is not None) else None
    diff_fwd_pe_soxx = (lseg_fwd_pe - soxx_fwd_pe) if (lseg_fwd_pe is not None and soxx_fwd_pe is not None) else None
    
    # 5. Append or Update in CSV
    new_data = {
        'Date': run_date,
        'LSEG_Price': lseg_price,
        'LSEG_PE': lseg_pe,
        'LSEG_FwdPE': lseg_fwd_pe,
        'YF_Index_Price': yf_index_price,
        'YF_PE': yf_pe,
        'YF_FwdPE': yf_fwd_pe,
        'SOXX_AsOfDate': soxx_date,
        'SOXX_PE': soxx_pe,
        'SOXX_FwdPE': soxx_fwd_pe,
        'Diff_PE_YF': diff_pe_yf,
        'Diff_FwdPE_YF': diff_fwd_pe_yf,
        'Ratio_PE_YF': ratio_pe_yf,
        'Ratio_FwdPE_YF': ratio_fwd_pe_yf,
        'Diff_PE_SOXX': diff_pe_soxx,
        'Diff_FwdPE_SOXX': diff_fwd_pe_soxx,
        'Missing_YF_Trail_Count': len(missing_t),
        'Missing_YF_Fwd_Count': len(missing_f),
        'Last_Updated_Time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    
    if os.path.exists(CSV_FILE):
        print(f"\n[CSV] Loading existing log file: {CSV_FILE}")
        df_log = pd.read_csv(CSV_FILE)
        # Rename Invesco columns if they exist
        if 'Invesco_PE' in df_log.columns:
            print("[CSV] Migrating old Invesco columns to SOXX columns...")
            df_log = df_log.rename(columns={
                'Invesco_AsOfDate': 'SOXX_AsOfDate',
                'Invesco_PE': 'SOXX_PE',
                'Invesco_FwdPE': 'SOXX_FwdPE',
                'Diff_PE_Invesco': 'Diff_PE_SOXX',
                'Diff_FwdPE_Invesco': 'Diff_FwdPE_SOXX'
            })
            
        # Check if row for today already exists
        if run_date in df_log['Date'].values:
            print(f"[CSV] Row for {run_date} already exists. Overwriting with new values.")
            idx = df_log[df_log['Date'] == run_date].index[0]
            for col, val in new_data.items():
                df_log.at[idx, col] = val
        else:
            print(f"[CSV] Appending new row for {run_date}.")
            df_new_row = pd.DataFrame([new_data])
            df_log = pd.concat([df_log, df_new_row], ignore_index=True)
    else:
        print(f"\n[CSV] Creating new log file: {CSV_FILE}")
        df_log = pd.DataFrame([new_data])
        
    df_log.to_csv(CSV_FILE, index=False)
    print(f"[CSV] Successfully saved log file to {os.path.abspath(CSV_FILE)}")
    
    try:
        save_to_excel_preserve_sheets(df_log, XLSX_FILE)
    except Exception as ex:
        print(f"[Excel] Error saving Excel log file: {ex}")
    
    # 6. Display comparison table
    print("\n" + "=" * 70)
    print("  SUMMARY COMPARISON TABLE")
    print("=" * 70)
    print(f"{'Source':<15} | {'Trailing PE':>12} | {'Forward PE':>12} | {'As-of Date':>12}")
    print("-" * 70)
    
    lseg_pe_str = f"{lseg_pe:.2f}x" if lseg_pe is not None else "N/A"
    lseg_fwd_str = f"{lseg_fwd_pe:.2f}x" if lseg_fwd_pe is not None else "N/A"
    print(f"{'LSEG Index':<15} | {lseg_pe_str:>12} | {lseg_fwd_str:>12} | {run_date:>12}")
    
    yf_pe_str = f"{yf_pe:.2f}x" if yf_pe is not None else "N/A"
    yf_fwd_str = f"{yf_fwd_pe:.2f}x" if yf_fwd_pe is not None else "N/A"
    print(f"{'yfinance (BU)':<15} | {yf_pe_str:>12} | {yf_fwd_str:>12} | {run_date:>12}")
    
    soxx_pe_str = f"{soxx_pe:.2f}x" if soxx_pe is not None else "N/A"
    soxx_fwd_str = f"{soxx_fwd_pe:.2f}x" if soxx_fwd_pe is not None else "N/A"
    soxx_date_str = soxx_date if soxx_date is not None else "N/A"
    print(f"{'iShares SOXX':<15} | {soxx_pe_str:>12} | {soxx_fwd_str:>12} | {soxx_date_str:>12}")
    
    print("-" * 70)
    ratio_pe_str = f"{ratio_pe_yf:.3f}" if ratio_pe_yf is not None else "N/A"
    ratio_fwd_str = f"{ratio_fwd_pe_yf:.3f}" if ratio_fwd_pe_yf is not None else "N/A"
    print(f"{'Ratio (LSEG/YF)':<15} | {ratio_pe_str:>12} | {ratio_fwd_str:>12} | {'-':>12}")
    
    diff_pe_str = f"{diff_pe_yf:+.2f}x" if diff_pe_yf is not None else "N/A"
    diff_fwd_str = f"{diff_fwd_pe_yf:+.2f}x" if diff_fwd_pe_yf is not None else "N/A"
    print(f"{'Diff (LSEG-YF)':<15} | {diff_pe_str:>12} | {diff_fwd_str:>12} | {'-':>12}")
    print("=" * 70)
    
    if len(missing_t) > 0 or len(missing_f) > 0:
        print(f"Note: yfinance had missing data. Trail missing: {len(missing_t)}, Fwd missing: {len(missing_f)}")
    print("Log process completed successfully.\n")

if __name__ == "__main__":
    main()
